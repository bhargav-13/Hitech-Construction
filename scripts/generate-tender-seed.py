# Regenerates src/lib/tenderSeed.ts from the client's tender-analysis workbook.
#
#   python scripts/generate-tender-seed.py "path/to/Tender Analysis.xlsx"
#
# Requires: pip install openpyxl. Every column of every sheet is represented; SORTING is capped
# (see CAPS) to keep the shipped seed small. Re-run after the workbook changes, then commit the seed.
#
# Beyond straight extraction this also NORMALISES the messier columns, because the app reports on
# them and free text cannot be summed:
#   * EMD TYPE      -> emdMode (instrument) + emdState (paid/pending), which the sheet conflated
#   * SECURITY      -> securityType / securityAmount / additionalSecurity* / bgCharges
#   * VALIDITY      -> validityDays,  DURATION -> durationMonths
#   * PRIORITY      -> HIGH/MEDIUM/LOW (the sheet contains the typo "Meduim")
#   * PRE BID INFO  -> preBidDate, so pre-bid meetings can reach the calendar
# The raw text is kept alongside every normalised field so nothing from the client is lost.
import sys, io, json, datetime, os, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import warnings; warnings.filterwarnings("ignore")
import openpyxl
DEFAULT_XLSX = r"C:\Users\bharg\Downloads\🏗️ Construction & 📑 GeM Tender Analysis 📊.xlsx"
p = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
wb = openpyxl.load_workbook(p, data_only=True)

def clean(v):
    if v is None: return None
    if isinstance(v, bool): return v
    if isinstance(v, (datetime.datetime, datetime.date)): return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer(): return int(v)
    if isinstance(v, str):
        s = v.replace("\r", " ").replace("\n", " ").strip()
        return s if s not in ("", "-") else None
    return v

def num(v):
    v = clean(v)
    if v is None: return None
    if isinstance(v, (int, float)): return v
    try: return float(str(v).replace(",", "").split("/")[0])
    except: return None

def sheet_rows(name, header_row):
    ws = wb[name]
    hdr = [clean(c) for c in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    out = []
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(c not in (None, '', '-') for c in r): continue
        out.append(list(r))
    return hdr, out

idx = 0
def nid():
    global idx; idx += 1; return "tnd-%03d" % idx

STATUS_MAP = {
    "Awarded (Win)": ("WON", "WON"), "Completed": ("WON", "COMPLETED"), "Lost": ("LOST", "LOST"),
    "Technically Disqualified": ("LOST", "TECH_DISQUALIFIED"), "Cancelled": ("LOST", "CANCELLED"),
    "Retendered": ("APPLIED", "RETENDERED"), "Submitted / Under Review": ("APPLIED", "SUBMITTED"),
    "Technical Opened": ("APPLIED", "TECH_OPENED"), "Technically Qualified": ("APPLIED", "TECH_QUALIFIED"),
}

# Only reasons the sheet actually implies. A plain "Lost" stays unspecified rather than inventing a
# cause — the app prompts the user for it, and "Unspecified" on the dashboard is the honest answer.
LOSS_REASON_FROM_STATUS = {
    "TECH_DISQUALIFIED": "TECHNICALLY_DISQUALIFIED",
    "CANCELLED": "CANCELLED_BY_DEPT",
    "RETENDERED": "RETENDERED",
}

def col(hdr, row, name):
    try: return clean(row[hdr.index(name)])
    except ValueError: return None
def coln(hdr, row, name):
    try: return num(row[hdr.index(name)])
    except ValueError: return None

# ---------- normalisers (kept in step with src/lib/tenderHelpers.ts) ----------

def emd_type(v):
    """EMD TYPE mixed instrument and payment status; split them."""
    s = re.sub(r"[.\s]", "", str(v or "")).upper()
    if not s or s == "-": return (None, None)
    if "EXEMPT" in s or "NIL" in s: return ("EXEMPT", "PAID")
    if "FDR" in s: return ("FDR", "PAID")
    if "BG" in s: return ("BG", "PAID")
    if "DD" in s or "DEMANDDRAFT" in s: return ("DD", "PAID")
    if "ONLINE" in s or "NEFT" in s or "RTGS" in s: return ("ONLINE", "PAID")
    if "DONE" in s or "PAID" in s: return (None, "PAID")
    if "PENDING" in s: return (None, "PENDING")
    return (None, None)

def sec_type(s):
    u = re.sub(r"[.\s]", "", s).upper()
    if "FDR" in u: return "FDR"
    if "BG" in u or "BANKGUARANTEE" in u: return "BG"
    if "CASH" in u or "ONLINE" in u: return "CASH"
    return None

def amount_in(s):
    m = re.search(r"([\d,]+(?:\.\d+)?)", s.replace(" ", ""))
    if not m: return None
    try: return float(m.group(1).replace(",", ""))
    except: return None

def security(v):
    """Parse 'S.D. Type : F.D.R.  Security : 97,200.00  Additional Security : ...' into fields."""
    out = {"securityType": None, "securityAmount": None, "additionalSecurityType": None,
           "additionalSecurityAmount": None, "bgCharges": None}
    if not v or str(v).strip() == "-": return out
    for part in re.split(r"(?=S\.?D\.?\s*Type|Security\s*:|Additional|B\.?G\.?\s*Charge)", str(v), flags=re.I):
        q = part.strip()
        if not q: continue
        if re.match(r"^S\.?D\.?\s*Type", q, re.I): out["securityType"] = sec_type(q)
        elif re.match(r"^Additional", q, re.I):
            out["additionalSecurityType"] = sec_type(q); out["additionalSecurityAmount"] = amount_in(q)
        elif re.match(r"^B\.?G\.?\s*Charge", q, re.I): out["bgCharges"] = amount_in(q)
        elif re.match(r"^Security", q, re.I): out["securityAmount"] = amount_in(q)
    return out

def duration_months(v):
    if v is None: return None
    if isinstance(v, (int, float)): return round(v)
    s = str(v).lower(); m = re.search(r"(\d+(?:\.\d+)?)", s)
    if not m: return None
    n = float(m.group(1))
    if "year" in s: return round(n * 12)
    if "day" in s: return round(n / 30)
    return round(n)

def validity_days(v):
    if v is None: return None
    if isinstance(v, (int, float)): return round(v)
    s = str(v).lower(); m = re.search(r"(\d+(?:\.\d+)?)", s)
    if not m: return None
    n = float(m.group(1))
    return round(n * 30) if "month" in s else round(n)

def priority(v):
    s = str(v or "").strip().upper()
    if not s: return None
    if s.startswith("H"): return "HIGH"
    if s.startswith("L"): return "LOW"
    if s.startswith("M"): return "MEDIUM"   # covers MEDIUM and the sheet's "MEDUIM"
    return None

def loose_date(v):
    if not v: return None
    s = str(v)
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m: return "%s-%s-%s" % m.groups()
    m = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})", s)
    if m: return "%s-%s-%s" % (m.group(3), m.group(2).zfill(2), m.group(1).zfill(2))
    return None

def add_months(iso, months):
    if not iso or months is None: return None
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", str(iso))
    if not m: return None
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    mo += months; y += (mo - 1) // 12; mo = (mo - 1) % 12 + 1
    import calendar
    return "%04d-%02d-%02d" % (y, mo, min(d, calendar.monthrange(y, mo)[1]))

def emd_fields(raw_type):
    mode, state = emd_type(raw_type)
    return {"emdType": clean(raw_type), "emdMode": mode, "emdState": state}

tenders = []
CAPS = {"SORTING": 120}
for sheet, stage in [("RESEARCH", "RESEARCH"), ("SORTING", "SORTING")]:
    hdr, rows = sheet_rows(sheet, 2)
    cap = CAPS.get(sheet)
    if cap: rows = rows[:cap]
    for row in rows:
        rec = {"id": nid(), "source": "PORTAL", "stage": stage,
            "department": col(hdr, row, "DEPARTMENT"), "tenderId": str(clean(col(hdr, row, "TENDER ID")) or ""),
            "nameOfWork": col(hdr, row, "NAME OF WORK"), "estimatedCost": coln(hdr, row, "ESTIMATED COST"),
            "deadline": col(hdr, row, "DEADLINE"), "duration": col(hdr, row, "DURATION"),
            "durationMonths": duration_months(col(hdr, row, "DURATION")),
            "preBidInfo": col(hdr, row, "PRE BID INFO"), "preBidDate": loose_date(col(hdr, row, "PRE BID INFO")),
            "validity": col(hdr, row, "VALIDITY"), "validityDays": validity_days(col(hdr, row, "VALIDITY")),
            "fee": coln(hdr, row, "FEE"), "emd": coln(hdr, row, "EMD"),
            "officeAddress": col(hdr, row, "OFFICE ADDRESS"), "hardcopyDue": col(hdr, row, "HARDCOPY DUE"),
            "techOpen": col(hdr, row, "TECH OPEN"), "priceOpen": col(hdr, row, "PRICE OPEN"),
            "pqCriteria": col(hdr, row, "PQ CRITERIA"), "classReq": col(hdr, row, "CLASS"), "gst": col(hdr, row, "GST"),
            "labTest": col(hdr, row, "LAB TEST"), "priceEscalation": col(hdr, row, "PRICE ESCALATION"),
            "depositDetails": col(hdr, row, "DEPOSIT DETAILS"), "dlp": col(hdr, row, "DLP"),
            "stageDocuments": col(hdr, row, "STAGE DOCUMENTS"), "viewDocuments": col(hdr, row, "VIEW DOCUMENTS"),
            "remarks": col(hdr, row, "REMARKS")}
        # Pipeline tenders have not paid EMD yet — that is what makes them "required" not "blocked".
        rec.update(emd_fields(col(hdr, row, "EMD TYPE")))
        if rec.get("emdState") is None: rec["emdState"] = "PENDING"
        tenders.append(rec)

for sheet, src in [("APPLIED", "PORTAL"), ("GEM APPLIED", "GEM")]:
    hdr, rows = sheet_rows(sheet, 2)
    for row in rows:
        st = col(hdr, row, "STATUS")
        stage, status = STATUS_MAP.get(st, ("APPLIED", "SUBMITTED")) if st else ("APPLIED", "SUBMITTED")
        rec = None
        for k in hdr:
            if isinstance(k, str) and k.startswith("RECEVIED"): rec = clean(row[hdr.index(k)])
        sd = col(hdr, row, "SECURITY DETAILS")
        t = {"id": nid(), "source": src, "stage": stage, "status": status, "statusLabel": st,
            "department": col(hdr, row, "DEPARTMENT"), "tenderId": str(clean(col(hdr, row, "TENDER ID")) or ""),
            "nameOfWork": col(hdr, row, "NAME OF WORK"), "location": col(hdr, row, "LOCATION"),
            "officeAddress": col(hdr, row, "OFFICE ADDRESS"), "duration": col(hdr, row, "DURATION"),
            "durationMonths": duration_months(col(hdr, row, "DURATION")), "dlp": col(hdr, row, "DLP"),
            "estimatedCost": coln(hdr, row, "ESTIMATED COST"), "variancePct": coln(hdr, row, "VAR.(%)"),
            "contractValue": coln(hdr, row, "CONTRACT VALUE"), "fee": coln(hdr, row, "FEE"),
            "emd": coln(hdr, row, "EMD"), "receivedStatus": rec,
            "dateOfReceived": col(hdr, row, "DATE OF RECEIVED"), "submissionDate": col(hdr, row, "DATE OF SUBMISSION"),
            "techOpen": col(hdr, row, "TECH OPEN"), "priceBidStage": col(hdr, row, "PRICE BID STAGE (COMPETITIVE REPORT)"),
            "validity": col(hdr, row, "VALIDITY"), "validityDays": validity_days(col(hdr, row, "VALIDITY")),
            "dueDate": col(hdr, row, "DUE DATE"), "securityDetails": sd, "firm": col(hdr, row, "FIRM"),
            "uploadedDocuments": col(hdr, row, "UPLOADED DOCUMENTS LIST"), "viewDocuments": col(hdr, row, "VIEW DOCUMENTS"),
            "remarks": col(hdr, row, "REMARKS")}
        t.update(emd_fields(col(hdr, row, "EMD TYPE")))
        t.update(security(sd))
        # A submitted bid always has its EMD furnished, whatever the sheet's EMD TYPE cell says.
        if t.get("emdState") is None: t["emdState"] = "PAID"
        # Money comes back once we lose; the sheet never recorded the refund, so it stays blocked
        # and shows up in the app's "recoverable EMD" list — which is the point of tracking it.
        if status in LOSS_REASON_FROM_STATUS: t["lossReason"] = LOSS_REASON_FROM_STATUS[status]
        tenders.append(t)

hdr, rows = sheet_rows("GEM SORTING", 2)
for row in rows:
    rec = {"id": nid(), "source": "GEM", "stage": "SORTING",
        "department": col(hdr, row, "DEPARTMENT"), "tenderId": str(clean(col(hdr, row, "TENDER ID")) or ""),
        "nameOfWork": col(hdr, row, "NAME OF WORK"), "estimatedCost": coln(hdr, row, "ESTIMATED COST"),
        "deadline": col(hdr, row, "DEADLINE"), "duration": col(hdr, row, "COMPLETION PERIOD"),
        "durationMonths": duration_months(col(hdr, row, "COMPLETION PERIOD")),
        "validity": col(hdr, row, "VALIDITY"), "validityDays": validity_days(col(hdr, row, "VALIDITY")),
        "emd": coln(hdr, row, "EMD"), "officeAddress": col(hdr, row, "OFFICE ADDRESS"),
        "openingDate": col(hdr, row, "OPENING DATE"),
        "gemCategory": col(hdr, row, "CATEGORY"), "msmeRelaxation": col(hdr, row, "MSME RELAXATION"),
        "experienceTurnover": col(hdr, row, "EXPERIENCE & TURNOVER"), "securityDetails": col(hdr, row, "SECURITY"),
        "dlp": col(hdr, row, "DLP"), "eligibilityStatus": col(hdr, row, "ELIGIBILITY STATUS"),
        "priority": priority(col(hdr, row, "PRIORITY")), "stageDocuments": col(hdr, row, "STAGE DOCUMENTS"),
        "viewDocuments": col(hdr, row, "VIEW DOCUMENTS"), "remarks": col(hdr, row, "REMARKS")}
    rec.update(security(col(hdr, row, "SECURITY")))
    rec["emdState"] = "PENDING"
    tenders.append(rec)

def b(v):
    if isinstance(v, bool): return v
    if v in (None, "", "-"): return False
    return str(v).strip().lower() in ("true", "yes", "done")

hdr, rows = sheet_rows("STATUS", 2)
def hcol(hdr, row, *names):
    for n in names:
        if n in hdr: return clean(row[hdr.index(n)])
    return None
milestones = []
for i, row in enumerate(rows, 1):
    milestones.append({"id": "mst-%03d" % i, "tenderId": str(clean(col(hdr, row, "TENDER ID")) or ""),
        "nameOfWork": col(hdr, row, "NAME OF WORK"),
        "analysis": b(col(hdr, row, "ANALYSIS")), "fee": b(col(hdr, row, "FEE")), "emd": b(col(hdr, row, "EMD")),
        "applied": b(col(hdr, row, "APPLIED")), "techBid": b(hcol(hdr, row, "TECH\nBID", "TECH BID")),
        "priceBid": b(hcol(hdr, row, "PRICE\nBID", "PRICE BID")), "loa": b(col(hdr, row, "LOA")),
        "security": b(col(hdr, row, "SECURITY")), "additionalSecurity": b(hcol(hdr, row, "ADDITIONAL\nSECURITY", "ADDITIONAL SECURITY")),
        "agreement": b(col(hdr, row, "AGREEMENT")), "workOrder": b(col(hdr, row, "WORK ORDER")),
        "workStartDate": col(hdr, row, "WORK START DATE"), "progress": col(hdr, row, "PROGRESS & STATUS")})

ws = wb["DOCUMENT STATUS"]
# Rows 2-4 are merged group / sub / SOFT-HARD label headers; real records have a TENDER ID in col 1.
# NOTE the RA-bill layout: the five SOFT flags sit in cols 15-19 and their HARD counterparts in
# 20-24 (not interleaved like every other category). Verified against the merged ranges P4:T4/U4:Y4.
drows = [list(r) for r in ws.iter_rows(min_row=4, values_only=True)
         if any(c not in (None, '', '-') for c in r) and clean(r[1]) is not None]
def pair(row, si, hi):
    return {"soft": bool(row[si]) if isinstance(row[si], bool) else False,
            "hard": bool(row[hi]) if isinstance(row[hi], bool) else False}
documents = []
for i, row in enumerate(drows, 1):
    documents.append({"id": "doc-%03d" % i, "tenderId": str(clean(row[1]) or ""), "nameOfWork": clean(row[2]),
        "competitiveReport": pair(row, 3, 4), "letterOfAcceptance": pair(row, 5, 6), "securityDeposit": pair(row, 7, 8),
        "stampDutyAgreement": pair(row, 9, 10), "workOrder": pair(row, 11, 12), "generalCorrespondence": pair(row, 13, 14),
        "raBills": [pair(row, 15 + j, 20 + j) for j in range(5)],
        "finalBill": pair(row, 25, 26), "form3a": pair(row, 27, 28),
        "progress": clean(row[29]), "viewDocuments": clean(row[31]) if len(row) > 31 else None})

# Index tenders by a squashed name so the sheets that only carry NAME OF WORK can still be linked.
def keyof(name):
    return re.sub(r"[^a-z0-9]", "", str(name or "").lower())[:48]
BY_NAME = {}
for t in tenders:
    k = keyof(t.get("nameOfWork"))
    if k and k not in BY_NAME: BY_NAME[k] = t

hdr, rows = sheet_rows("HARDCOPY TRACKER", 2)
hardcopy = []
for i, row in enumerate(rows, 1):
    name = col(hdr, row, "NAME OF WORK")
    linked = BY_NAME.get(keyof(name))
    hardcopy.append({"id": "hcp-%03d" % i, "date": col(hdr, row, "DATE"), "nameOfWork": name,
        # The source sheet has no TENDER ID column at all; resolve it here so the app can link rows.
        "tenderId": linked["tenderId"] if linked else None,
        "documentList": col(hdr, row, "DOCUMENT LIST"), "packedBy": col(hdr, row, "PACKED BY & MO."),
        "dispatchBy": col(hdr, row, "DISPATCH BY & MO."),
        "trackingNo": (lambda v: str(v) if v is not None else None)(col(hdr, row, "POST TRACKING NO.")),
        "arrived": hcol(hdr, row, "ARRIVED/\nNOT ARRIVED", "ARRIVED/ NOT ARRIVED"),
        "arrivedDate": col(hdr, row, "ARRIVED DATE"), "remarks": col(hdr, row, "REMARKS")})

hdr, rows = sheet_rows("MATERIAL PARTY DETAILS", 2)
materials = []
for i, row in enumerate(rows, 1):
    materials.append({"id": "mat-%03d" % i, "party": col(hdr, row, "PARTY DETAILS"),
        "manufacturerType": col(hdr, row, "TYPE OF MANUFACTURER"), "make": col(hdr, row, "MAKE"),
        "location": col(hdr, row, "LOCATION"), "contact": col(hdr, row, "CONTACT PERSON DETAILS"), "email": col(hdr, row, "EMAIL")})

# NOTE — PROJECTS WORK IN HAND, COMPLETED PROJECTS and PROJECT WISE RA BILLS are deliberately NOT
# seeded here. A won tender leaves this module and becomes a record in the Project module, which
# already exists and is backend-backed; duplicating those sheets inside Tender would fork the truth.
# Those three sheets belong to a Project-module importer.

def strip(o): return {k: v for k, v in o.items() if v is not None}
tenders = [strip(t) for t in tenders]; materials = [strip(m) for m in materials]
hardcopy = [strip(h) for h in hardcopy]; milestones = [strip(m) for m in milestones]

from collections import Counter
print("tenders", len(tenders), "stages", dict(Counter(t["stage"] for t in tenders)), "src", dict(Counter(t["source"] for t in tenders)))
print("emdMode", dict(Counter(t.get("emdMode") for t in tenders)), "emdState", dict(Counter(t.get("emdState") for t in tenders)))
print("security parsed", sum(1 for t in tenders if t.get("securityAmount")), "| lossReason", dict(Counter(t.get("lossReason") for t in tenders if t["stage"] == "LOST")))
print("milestones", len(milestones), "documents", len(documents), "hardcopy", len(hardcopy),
      "(linked", sum(1 for h in hardcopy if h.get("tenderId")), ") materials", len(materials))

def arr(name):
    return json.dumps(globals()[name], ensure_ascii=False)

ts = (
    "// AUTO-GENERATED from the client tender-analysis workbook via scripts/generate-tender-seed.py.\n"
    "// Every column from the source sheets is represented (SORTING capped to ~120 rows), with the\n"
    "// free-text EMD / security / validity / duration / priority columns normalised into typed fields.\n"
    "// Regenerate rather than hand-edit.\n"
    'import type { Tender, MaterialParty, TenderMilestones, TenderDocuments, HardcopyDispatch } from "./tenderTypes";\n\n'
    "export const TENDER_SEED: Tender[] = " + arr("tenders") + ";\n\n"
    "export const MILESTONE_SEED: TenderMilestones[] = " + arr("milestones") + ";\n\n"
    "export const DOCUMENT_SEED: TenderDocuments[] = " + arr("documents") + ";\n\n"
    "export const HARDCOPY_SEED: HardcopyDispatch[] = " + arr("hardcopy") + ";\n\n"
    "export const MATERIAL_SEED: MaterialParty[] = " + arr("materials") + ";\n"
)
out = os.path.join(os.path.dirname(__file__), "..", "src", "lib", "tenderSeed.ts")
with open(out, "w", encoding="utf-8") as f:
    f.write(ts)
print("wrote", os.path.abspath(out), round(len(ts) / 1024), "KB")
